import io

from openpyxl import load_workbook

from app.services.documents.chunking_service import TextChunk

_ROWS_PER_CHUNK = 50


def parse_xlsx(data: bytes) -> list[TextChunk]:
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    segments: list[TextChunk] = []

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c) if c is not None else "" for c in rows[0]]
        data_rows = rows[1:] or rows

        for start in range(0, len(data_rows), _ROWS_PER_CHUNK):
            batch = data_rows[start : start + _ROWS_PER_CHUNK]
            lines = []
            for row in batch:
                pairs = [
                    f"{header[i] if i < len(header) else f'col{i}'}={value}"
                    for i, value in enumerate(row)
                    if value is not None
                ]
                if pairs:
                    lines.append(", ".join(pairs))
            if not lines:
                continue
            end = start + len(batch)
            segments.append(
                TextChunk(
                    text="\n".join(lines),
                    section_title=f"sheet '{sheet.title}', rows {start + 1}-{end}",
                )
            )
    return segments
